import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

# Input/Output paths
JSON_FILE = 'users_full_data.json'
EXCEL_FILE = 'e-LMS_Users_Detailed.xlsx'

if not os.path.exists(JSON_FILE):
    print(f"Error: {JSON_FILE} not found")
    exit(1)

# Read Data
with open(JSON_FILE, 'r') as f:
    users = json.load(f)

# Create Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Detailed User List"

# Headers
headers = [
    "Username", "First Name", "Last Name", "Email", 
    "Role", "Password (Default)", "Courses Teaching", "Courses Enrolled"
]

ws.append(headers)

# Style Headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

# Write Data
for user in users:
    # Prepare row
    row = [
        user.get('username', ''),
        user.get('firstname', ''),
        user.get('lastname', ''),
        user.get('email', ''),
        user.get('role', ''),
        user.get('password', ''),
        user.get('courses_teaching', ''),
        user.get('courses_enrolled', '')
    ]
    ws.append(row)

# Auto-adjust column widths
for col_num, header in enumerate(headers, 1):
    column_letter = get_column_letter(col_num)
    max_length = 0
    
    # Check all cells in column
    for cell in ws[column_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
            
    # Cap width for long lists (Courses)
    if max_length > 50:
        adjusted_width = 50
    else:
        adjusted_width = (max_length + 2)
        
    ws.column_dimensions[column_letter].width = adjusted_width

# Wrap text for Course columns (G and H)
for row in ws.iter_rows(min_row=2, max_col=8, max_row=ws.max_row):
    # Teaching (Col 7) and Enrolled (Col 8)
    row[6].alignment = Alignment(wrap_text=True, vertical="top")
    row[7].alignment = Alignment(wrap_text=True, vertical="top")

# Save
try:
    wb.save(EXCEL_FILE)
    print(f"Successfully created: {os.path.abspath(EXCEL_FILE)}")
except Exception as e:
    print(f"Error saving file: {e}")
