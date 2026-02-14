import openpyxl
import json
from openpyxl.styles import Font, PatternFill

# Load JSON data
with open('users_export.json', 'r') as f:
    users = json.load(f)

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "e-LMS Users"

# Headers
headers = ['Username', 'First Name', 'Last Name', 'Email', 'Role(s)', 'City', 'Country', 'Password']
ws.append(headers)

# Style headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Add data
for user in users:
    row = [
        user['username'],
        user['firstname'],
        user['lastname'],
        user['email'],
        user['roles'],
        user['city'],
        user['country'],
        user['password_hint']
    ]
    ws.append(row)

# Auto-adjust column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save file
wb.save('e-LMS_Users.xlsx')
print("Created e-LMS_Users.xlsx successfully!")
